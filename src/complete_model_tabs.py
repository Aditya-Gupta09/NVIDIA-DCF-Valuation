"""
src/complete_model_tabs.py — Final Excel Tab Completion
=========================================================
Completes the three remaining manual Excel tasks from Step 1 / Step 3:

    1. Fill 98_Notes tab  (5 rows: version, author, date, sources, purpose)
    2. Create 07_Adjustments tab  (Arm termination + two other one-offs documented)
    3. Fix H93:L93 — link "Other, net" projection to 00_Assumptions
       by adding assumption row 39 and updating the formula references

Usage:
    python src/complete_model_tabs.py
    python src/complete_model_tabs.py --model path/to/model.xlsx --author "Your Name"
    python src/complete_model_tabs.py --dry-run

Output:
    Overwrites model in-place. Backup written first.

IMPORTANT:
    After running this script, open the model in Excel and press
    Ctrl+Alt+F9 to recalculate all formula cells, then save.
"""

import os
import sys
import shutil
import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── CONFIG ───────────────────────────────────────────────────────────────────

DEFAULT_MODEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "models", "Company_Valuation_Model.xlsx"
)

NAVY   = "1F3864"
BLUE   = "2E5FA3"
LBLUE  = "D6E4F7"
GREEN  = "E2EFDA"
YELLOW = "FFF2CC"
GREY   = "F2F2F2"
WHITE  = "FFFFFFFF"


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def style_cell(cell, bold=False, fill=None, font_color="000000",
               size=11, wrap=False, align_h="left"):
    cell.font      = Font(bold=bold, color=font_color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align_h, vertical="center",
                                wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)

def header_cell(ws, row, col, text, fill=BLUE, color="FFFFFF", width=None):
    cell = ws.cell(row=row, column=col, value=text)
    style_cell(cell, bold=True, fill=fill, font_color=color, size=11)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell

def data_cell(ws, row, col, text, fill=None, bold=False, wrap=False):
    cell = ws.cell(row=row, column=col, value=text)
    style_cell(cell, bold=bold, fill=fill, size=10, wrap=wrap)
    return cell

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ─── TASK 1: FILL 98_NOTES ────────────────────────────────────────────────────

def fill_98_notes(wb, author="[Your Name]"):
    """Populates the 98_Notes tab with version, author, date, sources, purpose."""

    ws = wb["98_Notes"]
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "NVIDIA Corporation — Financial Model Notes"
    style_cell(cell, bold=True, fill=NAVY, font_color="FFFFFF", size=14, align_h="center")
    ws.row_dimensions[1].height = 30

    # Sub-header
    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = "Model metadata, version history, and data source documentation"
    style_cell(sub, fill=LBLUE, font_color=BLUE, size=10, align_h="center")

    # Column headers (row 4)
    ws.row_dimensions[4].height = 22
    header_cell(ws, 4, 1, "Field",   fill=BLUE, width=28)
    header_cell(ws, 4, 2, "Value",   fill=BLUE, width=42)
    header_cell(ws, 4, 3, "Notes",   fill=BLUE, width=50)
    header_cell(ws, 4, 4, "Last Updated", fill=BLUE, width=16)

    # Data rows
    rows = [
        ("Version",        "1.0",                           "Initial release",                   "Oct 2025"),
        ("Author",         author,                          "Primary model builder",             "Oct 2025"),
        ("Valuation Date", "January 31, 2025",              "FY2025 close; market data Oct 2025","Oct 2025"),
        ("Data Sources",   "NVIDIA 10-K FY2025 (SEC EDGAR),"
                           " Yahoo Finance, Bloomberg Comps","All historical data from 10-K filings","Oct 2025"),
        ("Model Purpose",  "L3 DCF Valuation — NVIDIA Corporation (NVDA)",
                           "FCFF / Gordon Growth + Exit Multiple cross-check","Oct 2025"),
        ("Currency",       "USD millions (unless stated)",  "All $ figures in $M",               "Oct 2025"),
        ("Fiscal Year End","January 31",                    "NVIDIA FY ends late Jan/early Feb",  "Oct 2025"),
        ("Forecast Period","FY2026F – FY2030F (5 years)",   "Base case only",                    "Oct 2025"),
        ("WACC",           "12.91%",                        "Blume-adjusted beta; CAPM",         "Oct 2025"),
        ("Terminal Growth","4.00%",                         "Long-run GDP growth proxy",          "Oct 2025"),
    ]

    fills = [GREY if i % 2 == 0 else "FFFFFF" for i in range(len(rows))]

    for i, (field, value, note, updated) in enumerate(rows):
        r = i + 5
        ws.row_dimensions[r].height = 18
        f = fills[i]
        data_cell(ws, r, 1, field,   fill=f, bold=True)
        data_cell(ws, r, 2, value,   fill=f, wrap=True)
        data_cell(ws, r, 3, note,    fill=f, wrap=True)
        data_cell(ws, r, 4, updated, fill=f)
        for c in range(1, 5):
            ws.cell(r, c).border = thin_border()

    # Section 2: Change log
    r = len(rows) + 7
    ws.merge_cells(f"A{r}:D{r}")
    cell = ws.cell(r, 1, "Version History")
    style_cell(cell, bold=True, fill=NAVY, font_color="FFFFFF", size=12)
    ws.row_dimensions[r].height = 22

    r += 1
    for col, hdr in [(1,"Version"),(2,"Date"),(3,"Change"),(4,"Author")]:
        header_cell(ws, r, col, hdr, fill=BLUE)

    for version_row in [
        ("1.0", "Oct 2025", "Initial model build — historical IS/BS/CF, WACC, Comps, DCF", author),
        ("1.0", "Oct 2025", "Step 2: added named ranges, FCFF validation, data_loader.py", author),
        ("1.0", "Oct 2025", "Step 3: fixed Other net assumption, filled notes, adjustments tab", author),
    ]:
        r += 1
        ws.row_dimensions[r].height = 16
        for c, val in enumerate(version_row, 1):
            data_cell(ws, r, c, val, fill="FFFFFF")
            ws.cell(r, c).border = thin_border()

    print("  ✓  98_Notes tab populated")


# ─── TASK 2: CREATE 07_ADJUSTMENTS ───────────────────────────────────────────

def create_07_adjustments(wb):
    """Creates the 07_Adjustments tab documenting one-off model items."""

    # Insert after 06_Sensitivity (index 6), before 07_CommonSize
    # Find position
    sheet_names = wb.sheetnames
    insert_after = "06_Sensitivity"
    if "07_Adjustments" in sheet_names:
        print("  ─  07_Adjustments already exists — refreshing content")
        ws = wb["07_Adjustments"]
        # Clear existing
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        # Create after 06_Sensitivity
        idx = sheet_names.index(insert_after) + 1 if insert_after in sheet_names else 6
        ws = wb.create_sheet("07_Adjustments", idx)

    ws.sheet_view.showGridLines = False

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "NVIDIA Corporation — Non-Recurring Items & Model Adjustments"
    style_cell(cell, bold=True, fill=NAVY, font_color="FFFFFF", size=14, align_h="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = (
        "Documents all one-time, non-recurring, and adjusted items excluded from or adjusted "
        "within the DCF model. Per Phase A spec — all items with justifications."
    )
    style_cell(sub, fill=LBLUE, font_color=BLUE, size=10, align_h="center", wrap=True)
    ws.row_dimensions[2].height = 30

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {1: 6, 2: 32, 3: 12, 4: 14, 5: 20, 6: 45, 7: 30}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Section 1: One-time items ──────────────────────────────────────────
    ws.merge_cells("A4:G4")
    cell = ws["A4"]
    cell.value = "Section 1 — One-Time / Non-Recurring Items  (Historical)"
    style_cell(cell, bold=True, fill=BLUE, font_color="FFFFFF", size=12)
    ws.row_dimensions[4].height = 22

    # Headers row 5
    headers = ["#", "Item", "FY Year", "Amount ($M)", "P&L Line", "Justification / Treatment", "Source"]
    ws.row_dimensions[5].height = 20
    for c, h in enumerate(headers, 1):
        header_cell(ws, 5, c, h, fill=BLUE)

    # One-time items
    items = [
        (1, "Arm Holdings Acquisition Termination Fee",
         "FY2023", "1,350",
         "Acquisition termination cost (OPEX)",
         "One-time charge: NVIDIA terminated its $40B Arm acquisition in Feb 2022 "
         "due to regulatory opposition. Fee paid to SoftBank. "
         "TREATMENT: Included in FY2023 historical EBIT but EXCLUDED from "
         "FY2024F+ projections (zeroed in 04a row 88). Excluded from FCFF "
         "normalised EBIT in 05a row 9.",
         "NVIDIA 10-K FY2023, Note 2"),

        (2, "Deferred Income Tax Benefit (FY2023)",
         "FY2023", "(2,164)",
         "Income tax (CF Statement)",
         "Large non-cash deferred tax benefit in FY2023 drove negative effective tax rate (-4.5%). "
         "TREATMENT: Reflected in historical data only. "
         "Forward ETR normalised to 15% base case (00_Assumptions row 36).",
         "NVIDIA 10-K FY2023, Cash Flow"),

        (3, "Gains on Non-Marketable Equity Securities",
         "FY2024–FY2025", "238 / 1,030",
         "Other income (below EBIT)",
         "Non-cash gains on strategic equity investments (AI/startup portfolio). "
         "TREATMENT: Included in historical 'Other, net'. "
         "Not projected forward in base case — excluded from FCFF unlevered framework. "
         "FY2026F+ Other net assumption = $1,000M flat (conservative).",
         "NVIDIA 10-K FY2024-FY2025"),
    ]

    fills = [GREY, "FFFFFF", GREY]
    for i, (num, name, yr, amt, line, just, source) in enumerate(items):
        r = i + 6
        ws.row_dimensions[r].height = 60
        f = fills[i]
        for c, val in enumerate([str(num), name, yr, amt, line, just, source], 1):
            cell = ws.cell(r, c, val)
            style_cell(cell, fill=f, size=9 if c == 6 else 10, wrap=True)
            cell.border = thin_border()

    # ── Section 2: Model normalisation notes ──────────────────────────────
    r = len(items) + 8
    ws.merge_cells(f"A{r}:G{r}")
    cell = ws.cell(r, 1, "Section 2 — Model Normalisation Notes")
    style_cell(cell, bold=True, fill=BLUE, font_color="FFFFFF", size=12)
    ws.row_dimensions[r].height = 22

    r += 1
    for c, h in enumerate(["#", "Topic", "Normalisation Applied", "Rationale", "", "", ""], 1):
        if h:
            header_cell(ws, r, c, h, fill=BLUE)

    norm_items = [
        (1, "Stock-Based Compensation (SBC)",
         "Added back in CFO but NOT in FCFF (unlevered framework). "
         "SBC FY2025 = $4,737M (~3.6% revenue). Not projected separately — "
         "treated as operating cost already embedded in EBIT margins.",
         "Standard FCFF treatment: SBC is non-cash but real economic cost."),

        (2, "Capex as % of Revenue",
         "FY2026F: 3.0% of revenue (linked to 00_Assumptions!C42 / named range Capex_pct). "
         "Historical: FY2023=6.8%, FY2024=1.8%, FY2025=2.5%. "
         "3.0% forward reflects NVIDIA's asset-light fabless model.",
         "Fabless model = low capex intensity vs. foundries (TSMC capex ~25%+)."),

        (3, "D&A Projection",
         "FY2026F D&A = $3,046M (from 11_DA schedule, 5-yr avg useful life). "
         "Grows with capex additions. Linked directly from DA schedule.",
         "D&A must track capex spend; using fixed asset roll-forward is most accurate."),
    ]

    for i, (num, topic, normalisation, rationale) in enumerate(norm_items):
        r += 1
        ws.row_dimensions[r].height = 50
        f = GREY if i % 2 == 0 else "FFFFFF"
        ws.merge_cells(f"D{r}:G{r}")
        for c, val in enumerate([str(num), topic, normalisation, rationale], 1):
            cell = ws.cell(r, c, val)
            style_cell(cell, fill=f, size=9, wrap=True)
            cell.border = thin_border()

    print("  ✓  07_Adjustments tab created")


# ─── TASK 3: FIX H93:L93 — OTHER NET → 00_ASSUMPTIONS ───────────────────────

def fix_other_net_assumption(wb):
    """
    Adds 'Other income (net)' assumption to 00_Assumptions (row 39)
    and updates 04a_Projection_IS H93:L93 to reference it.

    Before: H93=1700, I93=1000, J93=1000, K93=1000, L93=1000 (hardcoded)
    After:  Row 39 in 00_Assumptions holds the base case values
            H93:L93 = formula referencing 00_Assumptions!$C$39 etc.
    """
    ws_assum = wb["00_Assumptions"]
    ws_proj  = wb["04a_Projection_IS"]

    # ── Add to 00_Assumptions row 39 ──────────────────────────────────────
    # Currently row 38 = "Assumed yield on cash", row 40 = section 5 header
    # Insert assumption at row 39 (fits between yield and section 5)

    # Check if already added
    existing = ws_assum.cell(39, 1).value
    if existing and "Other income" in str(existing):
        print("  ─  Other income (net) assumption already in 00_Assumptions row 39")
    else:
        # Write the assumption row
        ws_assum.cell(39, 1).value = "Other income (net) — Base case ($M)"
        ws_assum.cell(39, 1).font  = Font(name="Calibri", size=11)

        # Values: FY2026F=1700, FY2027F-FY2030F=1000
        # Store in columns C-G (matching the 00_Assumptions layout)
        for c, val in enumerate([1700, 1000, 1000, 1000, 1000], 3):
            cell = ws_assum.cell(39, c)
            cell.value = val
            cell.font  = Font(name="Calibri", size=11)
            cell.fill  = PatternFill("solid", fgColor="FFF2CC")  # yellow = assumption cell

        ws_assum.cell(39, 2).value = "Flat $1,000M from FY27 (conservative; excl. equity gains)"
        ws_assum.cell(39, 2).font  = Font(name="Calibri", size=10, italic=True, color="595959")

        print("  ✓  Added 'Other income (net)' to 00_Assumptions row 39")

    # ── Update 04a_Projection_IS H93:L93 to reference 00_Assumptions ─────
    # The projection columns are H=FY2026F, I=FY2027F, J=FY2028F, K=FY2029F, L=FY2030F
    # Map to 00_Assumptions columns C=FY2026F, D=FY2027F, E=FY2028F, F=FY2029F, G=FY2030F
    assum_cols = {8: "C", 9: "D", 10: "E", 11: "F", 12: "G"}

    for proj_col, assum_col in assum_cols.items():
        cell = ws_proj.cell(93, proj_col)
        old_val = cell.value
        new_formula = f"='00_Assumptions'!${assum_col}$39"
        cell.value = new_formula
        print(f"  ✓  04a_Projection_IS!{cell.column_letter}93: {old_val} → {new_formula}")

    print("  ✓  H93:L93 now linked to 00_Assumptions row 39")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Complete remaining Excel tabs: 98_Notes, 07_Adjustments, fix H93"
    )
    parser.add_argument("--model",   default=DEFAULT_MODEL_PATH)
    parser.add_argument("--author",  default="[Your Name]",
                        help="Your name for the 98_Notes author field")
    parser.add_argument("--dry-run", action="store_true",
                        help="List what would be done without writing")
    args = parser.parse_args()

    model_path = os.path.abspath(args.model)

    print(f"\n{'─'*64}")
    print(f"  complete_model_tabs.py — Final Tab Completion")
    print(f"{'─'*64}")
    print(f"\n  Model  : {model_path}")
    print(f"  Author : {args.author}")
    print(f"  Mode   : {'DRY RUN' if args.dry_run else 'WRITE'}\n")

    if not os.path.exists(model_path):
        print(f"  ERROR: Model not found: {model_path}\n")
        sys.exit(1)

    if args.dry_run:
        print("  Would perform:")
        print("    1. Fill 98_Notes tab (10 metadata rows + version history)")
        print("    2. Create 07_Adjustments tab (3 one-time items + 3 normalisation notes)")
        print("    3. Fix H93:L93 in 04a_Projection_IS → link to 00_Assumptions row 39")
        print("\n  DRY RUN — no changes written.\n")
        return 0

    # Backup
    backup = model_path.replace(".xlsx", ".BACKUP.xlsx")
    shutil.copy2(model_path, backup)
    print(f"  Backup: {backup}\n")

    # Load (data_only=False to write)
    wb = openpyxl.load_workbook(model_path, data_only=False)

    print("  TASK 1: 98_Notes")
    fill_98_notes(wb, author=args.author)

    print("\n  TASK 2: 07_Adjustments")
    create_07_adjustments(wb)

    print("\n  TASK 3: Fix H93:L93 (Other, net)")
    fix_other_net_assumption(wb)

    # Save
    wb.save(model_path)
    print(f"\n  Saved: {model_path}")

    print(f"\n{'─'*64}")
    print(f"  All 3 tasks complete.")
    print(f"\n  ⚠  NEXT STEP: Open the model in Excel →")
    print(f"     Press Ctrl+Alt+F9  →  Save")
    print(f"     This restores all formula-calculated values.")
    print(f"{'─'*64}\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
