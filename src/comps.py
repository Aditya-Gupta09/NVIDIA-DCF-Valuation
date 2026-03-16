"""
src/comps.py — Comparable Company Analysis
============================================
Reads peer company data from the Excel model's Comparable Analysis sheet,
calculates trading multiples for each peer, derives implied NVIDIA valuations
using median / mean / percentile multiples, and writes structured results to
datasets/processed/.

Peer universe (5 companies, excluding NVIDIA):
    AMD, Intel, Qualcomm (QCOM), Broadcom (AVGO), TSMC

Multiples calculated per peer:
    EV / Revenue
    EV / EBITDA
    P / E  (Price / EPS — excludes negative earners)

NVIDIA implied price derived three ways:
    EV multiple → Implied EV → minus Net Debt → Implied Market Cap → ÷ Shares
    P/E multiple → directly × NVIDIA EPS

Statistics computed:
    High, 75th percentile, Mean, Median, 25th percentile, Low

Usage:
    python src/comps.py
    python src/comps.py --model path/to/model.xlsx
    python src/comps.py --verbose

Output:
    datasets/processed/comps_results.json
"""

import os
import sys
import json
import argparse
import statistics
from datetime import datetime

import openpyxl


# ─── CONFIGURATION ────────────────────────────────────────────────────────────

DEFAULT_MODEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "models", "Company_Valuation_Model.xlsx"
)

OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "..", "datasets", "processed")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "comps_results.json")

SHEET_NAME  = "Comparable Analysis"

# Cell map: Comparable Analysis sheet row/col layout
# Row 6 = NVIDIA (subject), Rows 7-11 = peers
SUBJECT_ROW = 6
PEER_ROWS   = [7, 8, 9, 10, 11]

COL = {
    "company":    1,    # A
    "price":      2,    # B — share price ($)
    "eps":        3,    # C — earnings per share ($)
    "shares_b":   4,    # D — shares outstanding (billions) — NVIDIA only
    "mkt_cap_b":  5,    # E — market cap ($B)
    "net_debt_b": 6,    # F — net debt ($B) — NVIDIA only
    "ev_b":       7,    # G — enterprise value ($B)
    "revenue_b":  9,    # I — revenue ($B)
    "ebitda_b":   10,   # J — EBITDA ($B)
    "net_income_b": 11, # K — net income ($B)
}

# Rounding precision for display
PRICE_DP    = 2
MULTIPLE_DP = 2
PCT_DP      = 1


# ─── DATA LOADING ─────────────────────────────────────────────────────────────

def load_comps_data(wb):
    """
    Reads subject (NVIDIA) and peer company data from the Comparable Analysis sheet.
    Returns (subject_dict, list_of_peer_dicts).
    """
    ws = wb[SHEET_NAME]

    def read_row(row_idx):
        rec = {}
        for field, col_idx in COL.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            rec[field] = float(val) if isinstance(val, (int, float)) else None
        return rec

    subject = read_row(SUBJECT_ROW)

    # Subject-specific: company name from cell A6
    subject["company"] = ws.cell(row=SUBJECT_ROW, column=COL["company"]).value or "NVIDIA"

    peers = []
    for row in PEER_ROWS:
        peer = read_row(row)
        peer["company"] = ws.cell(row=row, column=COL["company"]).value or f"Peer_row{row}"
        peers.append(peer)

    return subject, peers


# ─── MULTIPLE CALCULATION ─────────────────────────────────────────────────────

def compute_multiples(peer):
    """
    Computes EV/Revenue, EV/EBITDA, and P/E for a single peer.
    Returns dict of {multiple_name: value_or_None}.
    Excludes a multiple if any required input is missing, zero, or produces
    an economically nonsensical result (e.g. negative earnings for P/E).
    """
    multiples = {}

    # EV / Revenue
    if peer["ev_b"] and peer["revenue_b"] and peer["revenue_b"] > 0:
        multiples["ev_revenue"] = peer["ev_b"] / peer["revenue_b"]
    else:
        multiples["ev_revenue"] = None

    # EV / EBITDA
    if peer["ev_b"] and peer["ebitda_b"] and peer["ebitda_b"] > 0:
        multiples["ev_ebitda"] = peer["ev_b"] / peer["ebitda_b"]
    else:
        multiples["ev_ebitda"] = None

    # P / E  — exclude negative earners (meaningless multiple)
    if peer["price"] and peer["eps"] and peer["eps"] > 0:
        multiples["pe"] = peer["price"] / peer["eps"]
    else:
        multiples["pe"] = None

    return multiples


# ─── STATISTICS ───────────────────────────────────────────────────────────────

def compute_stats(values):
    """
    Given a list of floats (Nones already filtered), computes:
    high, p75, mean, median, p25, low.
    Returns a dict.
    """
    clean = sorted([v for v in values if v is not None])
    if not clean:
        return {k: None for k in ["high", "p75", "mean", "median", "p25", "low", "n"]}

    n = len(clean)

    def percentile(data, p):
        # Linear interpolation, same as Excel's PERCENTILE.INC
        idx = (p / 100) * (len(data) - 1)
        lo  = int(idx)
        hi  = lo + 1
        if hi >= len(data):
            return data[-1]
        frac = idx - lo
        return data[lo] + frac * (data[hi] - data[lo])

    return {
        "high":   max(clean),
        "p75":    percentile(clean, 75),
        "mean":   sum(clean) / n,
        "median": statistics.median(clean),
        "p25":    percentile(clean, 25),
        "low":    min(clean),
        "n":      n,
    }


# ─── IMPLIED PRICE CALCULATION ────────────────────────────────────────────────

def implied_price_from_ev_multiple(multiple, subject_metric_b, net_debt_b, shares_b):
    """
    Derives NVIDIA implied share price from an EV-based multiple.

    Formula:
        Implied EV         = multiple × subject financial metric
        Implied Equity Val = Implied EV − Net Debt
        Implied Price      = Implied Equity Val / Shares Outstanding
    """
    if multiple is None or subject_metric_b is None or shares_b is None:
        return None
    implied_ev     = multiple * subject_metric_b
    implied_equity = implied_ev - net_debt_b          # net_debt is negative for NVIDIA (net cash)
    implied_price  = (implied_equity * 1e9) / (shares_b * 1e9)  # both in billions, cancel out
    return implied_price


def implied_price_from_pe(pe_multiple, subject_eps):
    """
    Derives NVIDIA implied share price from P/E multiple.
    Formula: Implied Price = P/E × EPS
    """
    if pe_multiple is None or subject_eps is None:
        return None
    return pe_multiple * subject_eps


# ─── PREMIUM / DISCOUNT ───────────────────────────────────────────────────────

def premium_discount(current_price, implied_price):
    """
    Returns upside/downside of implied price vs current market price.
    Positive = upside, Negative = discount to current price.
    """
    if implied_price is None or current_price is None or current_price == 0:
        return None
    return (implied_price / current_price - 1)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Comparable company analysis")
    parser.add_argument(
        "--model", default=DEFAULT_MODEL_PATH,
        help=f"Path to Excel model (default: models/Company_Valuation_Model.xlsx)"
    )
    parser.add_argument("--verbose", action="store_true")
    args = parser.parse_args()

    model_path = os.path.abspath(args.model)

    print(f"\n{'─'*64}")
    print(f"  comps.py — Comparable Company Analysis")
    print(f"{'─'*64}")
    print(f"\n  Model : {model_path}\n")

    # ── Validate file ──────────────────────────────────────────────────────
    if not os.path.exists(model_path):
        print(f"  ERROR: Model file not found: {model_path}\n")
        sys.exit(1)

    # ── Load ──────────────────────────────────────────────────────────────
    print(f"  [1/5] Loading model...")
    wb      = openpyxl.load_workbook(model_path, data_only=True)
    subject, peers = load_comps_data(wb)

    current_price = subject["price"]
    shares_b      = subject["shares_b"]
    net_debt_b    = subject["net_debt_b"]  # negative = net cash position
    nvidia_rev    = subject["revenue_b"]
    nvidia_ebitda = subject["ebitda_b"]
    nvidia_eps    = subject["eps"]

    print(f"  [2/5] Subject: {subject['company']}  |  "
          f"Current price: ${current_price}  |  Shares: {shares_b}B  |  "
          f"Net Debt: ${net_debt_b}B\n")

    # ── Compute peer multiples ─────────────────────────────────────────────
    print(f"  [3/5] Computing peer multiples...")
    peer_results = []
    for peer in peers:
        mults = compute_multiples(peer)
        peer_results.append({**peer, **mults})

    # ── Statistics ────────────────────────────────────────────────────────
    print(f"  [4/5] Computing statistics and implied prices...")
    ev_rev_vals  = [p["ev_revenue"] for p in peer_results]
    ev_ebitda_vals = [p["ev_ebitda"] for p in peer_results]
    pe_vals      = [p["pe"] for p in peer_results]

    stats = {
        "ev_revenue":  compute_stats(ev_rev_vals),
        "ev_ebitda":   compute_stats(ev_ebitda_vals),
        "pe":          compute_stats(pe_vals),
    }

    # ── Implied prices at each statistic level ─────────────────────────────
    stat_labels  = ["high", "p75", "mean", "median", "p25", "low"]
    implied = {}
    for stat in stat_labels:
        implied[stat] = {
            "ev_revenue": implied_price_from_ev_multiple(
                stats["ev_revenue"][stat], nvidia_rev, net_debt_b, shares_b
            ),
            "ev_ebitda": implied_price_from_ev_multiple(
                stats["ev_ebitda"][stat], nvidia_ebitda, net_debt_b, shares_b
            ),
            "pe": implied_price_from_pe(
                stats["pe"][stat], nvidia_eps
            ),
        }

    # ── Print results ──────────────────────────────────────────────────────
    print(f"\n{'─'*64}")
    print(f"  PEER TRADING MULTIPLES")
    print(f"{'─'*64}")
    print(f"\n  {'Company':<22} {'EV/Revenue':>11} {'EV/EBITDA':>11} {'P/E':>8}")
    print(f"  {'─'*20} {'─'*11} {'─'*11} {'─'*8}")

    for p in peer_results:
        ev_r  = f"{p['ev_revenue']:.1f}x"  if p["ev_revenue"] else "  n/m"
        ev_e  = f"{p['ev_ebitda']:.1f}x"   if p["ev_ebitda"]  else "  n/m"
        pe    = f"{p['pe']:.1f}x"          if p["pe"]         else "  n/m"
        print(f"  {p['company']:<22} {ev_r:>11} {ev_e:>11} {pe:>8}")

    print(f"\n{'─'*64}")
    print(f"  MULTIPLE STATISTICS  (n={stats['ev_revenue']['n']} peers for EV multiples)")
    print(f"{'─'*64}")
    print(f"\n  {'Statistic':<14} {'EV/Revenue':>11} {'EV/EBITDA':>11} {'P/E':>8}  "
          f"  {'Implied via EV/Rev':>19} {'EV/EBITDA':>10} {'P/E':>8}")
    print(f"  {'─'*12} {'─'*11} {'─'*11} {'─'*8}  {'─'*19} {'─'*10} {'─'*8}")

    labels_display = {
        "high":   "High",
        "p75":    "75th Pctile",
        "mean":   "Mean",
        "median": "Median",
        "p25":    "25th Pctile",
        "low":    "Low",
    }

    for stat in stat_labels:
        ev_r   = stats["ev_revenue"][stat]
        ev_e   = stats["ev_ebitda"][stat]
        pe     = stats["pe"][stat]
        i_evr  = implied[stat]["ev_revenue"]
        i_eve  = implied[stat]["ev_ebitda"]
        i_pe   = implied[stat]["pe"]

        ev_r_s  = f"{ev_r:.1f}x"    if ev_r  is not None else "   n/m"
        ev_e_s  = f"{ev_e:.1f}x"    if ev_e  is not None else "   n/m"
        pe_s    = f"{pe:.1f}x"      if pe    is not None else "   n/m"
        i_evr_s = f"${i_evr:.2f}"   if i_evr is not None else "    n/m"
        i_eve_s = f"${i_eve:.2f}"   if i_eve is not None else "    n/m"
        i_pe_s  = f"${i_pe:.2f}"    if i_pe  is not None else "    n/m"

        row_label = labels_display[stat]
        is_median = stat == "median"
        prefix = "→ " if is_median else "  "
        print(f"{prefix} {row_label:<14} {ev_r_s:>11} {ev_e_s:>11} {pe_s:>8}  "
              f"  {i_evr_s:>19} {i_eve_s:>10} {i_pe_s:>8}")

    # ── Median focus summary ───────────────────────────────────────────────
    med_evr = implied["median"]["ev_revenue"]
    med_eve = implied["median"]["ev_ebitda"]
    med_pe  = implied["median"]["pe"]

    updown_evr = premium_discount(current_price, med_evr)
    updown_eve = premium_discount(current_price, med_eve)
    updown_pe  = premium_discount(current_price, med_pe)

    def fmt_ud(v):
        if v is None: return "n/m"
        return f"+{v*100:.1f}%" if v >= 0 else f"{v*100:.1f}%"

    print(f"\n{'─'*64}")
    print(f"  MEDIAN IMPLIED PRICES vs CURRENT ${current_price:.2f}")
    print(f"{'─'*64}")
    print(f"\n  EV/Revenue  →  ${med_evr:.2f}  ({fmt_ud(updown_evr)} vs current price)")
    print(f"  EV/EBITDA   →  ${med_eve:.2f}  ({fmt_ud(updown_eve)} vs current price)")
    print(f"  P/E         →  ${med_pe:.2f}   ({fmt_ud(updown_pe)} vs current price)")
    print()

    # ── Build output JSON ──────────────────────────────────────────────────
    def _round(v, dp=4):
        return round(v, dp) if v is not None else None

    output = {
        "meta": {
            "script":        "src/comps.py",
            "model_file":    os.path.basename(model_path),
            "run_timestamp": datetime.now().isoformat(),
            "peer_universe": [p["company"] for p in peer_results],
            "subject":       subject["company"],
        },
        "subject_inputs": {
            "current_price":     current_price,
            "eps":               nvidia_eps,
            "shares_outstanding_b": shares_b,
            "net_debt_b":        net_debt_b,
            "revenue_b":         nvidia_rev,
            "ebitda_b":          nvidia_ebitda,
            "net_income_b":      subject["net_income_b"],
        },
        "peer_multiples": [
            {
                "company":    p["company"],
                "ev_b":       p["ev_b"],
                "revenue_b":  p["revenue_b"],
                "ebitda_b":   p["ebitda_b"],
                "eps":        p["eps"],
                "price":      p["price"],
                "ev_revenue": _round(p["ev_revenue"]),
                "ev_ebitda":  _round(p["ev_ebitda"]),
                "pe":         _round(p["pe"]),
            }
            for p in peer_results
        ],
        "multiple_statistics": {
            metric: {
                stat: _round(stats[metric][stat])
                for stat in stat_labels + ["n"]
            }
            for metric in ["ev_revenue", "ev_ebitda", "pe"]
        },
        "implied_prices": {
            stat: {
                "ev_revenue": _round(implied[stat]["ev_revenue"]),
                "ev_ebitda":  _round(implied[stat]["ev_ebitda"]),
                "pe":         _round(implied[stat]["pe"]),
            }
            for stat in stat_labels
        },
        "summary": {
            "current_price":            current_price,
            "median_implied_ev_revenue": _round(med_evr),
            "median_implied_ev_ebitda":  _round(med_eve),
            "median_implied_pe":         _round(med_pe),
            "updown_vs_current_ev_revenue": _round(updown_evr, 4),
            "updown_vs_current_ev_ebitda":  _round(updown_eve, 4),
            "updown_vs_current_pe":         _round(updown_pe, 4),
        }
    }

    # ── Write JSON ─────────────────────────────────────────────────────────
    print(f"  [5/5] Writing output...")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2)

    print(f"\n{'─'*64}")
    print(f"  Output written: {os.path.abspath(OUTPUT_FILE)}")
    print(f"{'─'*64}\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
