"""
export_datasets.py
===================
Exports three CSV files from Company_Valuation_Model.xlsx:

  datasets/raw/nvidia_historical_financials.csv   — IS + BS + CF (6 yrs)
  datasets/raw/nvidia_market_data.csv             — peer comps + market data
  datasets/processed/cleaned_financials.csv       — normalized per Phase A spec

Usage:
    python export_datasets.py
    python export_datasets.py --model path/to/model.xlsx
"""

import os
import csv
import argparse
from datetime import datetime
import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
DEFAULT_MODEL = os.path.join(
    os.path.dirname(__file__), "models", "Company_Valuation_Model.xlsx"
)
RAW_DIR       = os.path.join(os.path.dirname(__file__), "datasets", "raw")
PROCESSED_DIR = os.path.join(os.path.dirname(__file__), "datasets", "processed")

TICKER = "NVDA"


# ── Helpers ───────────────────────────────────────────────────────────────────

def v(ws, row, col):
    """Read cell value; return None if missing or non-numeric label."""
    val = ws.cell(row=row, column=col).value
    if isinstance(val, (int, float)):
        return val
    return None


def safe(val, default=None):
    return val if val is not None else default


def pct(numerator, denominator):
    if denominator and denominator != 0:
        return round(numerator / denominator, 6)
    return None


# ── 1. Historical Financials (raw) ────────────────────────────────────────────

def export_historical_financials(wb, out_path):
    """
    Exports IS (FY2020–2025), BS (FY2021–2025), CF (FY2023–2025)
    into a single long-format CSV, one row per year per statement.
    """

    ws_is = wb["01_Historical_IS"]
    ws_bs = wb["02_Historical_BS"]
    ws_cf = wb["03_Historical_CF"]

    # ── IS data: row 2 = years [2020,2021,2022,2023,2024,2025] in cols B–G ──
    is_years  = [int(ws_is.cell(row=2, column=c).value)
                 for c in range(2, 8)
                 if isinstance(ws_is.cell(row=2, column=c).value, int)]

    # Map: field -> row in IS sheet
    IS_ROWS = {
        "revenue":           3,
        "cogs":              4,
        "gross_profit":      5,
        "rd_expense":        8,
        "sga_expense":       9,
        "acq_termination":   10,
        "total_opex":        11,
        "ebit":              12,
        "da":                13,
        "interest_income":   15,
        "interest_expense":  16,
        "other_net":         17,
        "ebt":               19,
        "income_tax":        20,
        "net_income":        21,
    }

    # IS data: col B=FY2020 ... col G=FY2025
    is_data = {}
    for year_idx, year in enumerate(is_years):
        col = year_idx + 2
        row_dict = {"year": year, "ticker": TICKER, "statement": "IS"}
        for field, row in IS_ROWS.items():
            row_dict[field] = v(ws_is, row, col)
        is_data[year] = row_dict

    # ── BS data: row 2 = ['Fiscal Year','2021','2022','2023','2024','2025'] ──
    bs_years = []
    for c in range(2, 8):
        val = ws_bs.cell(row=2, column=c).value
        if isinstance(val, str) and val.strip().isdigit():
            bs_years.append((int(val.strip()), c))
        elif isinstance(val, int):
            bs_years.append((val, c))

    BS_ROWS = {
        "cash_equivalents":          4,
        "short_term_investments":     5,
        "cash_and_st_investments":    6,
        "receivables":                7,
        "inventory":                  8,
        "prepaid_expenses":           9,
        "total_current_assets":      10,
        "ppe_net":                   11,
        "operating_lease_assets":    12,
        "goodwill":                  13,
        "intangible_assets":         14,
        "lt_deferred_tax_assets":    15,
        "other_assets":              16,
        "total_assets":              17,
        "accounts_payable":          18,
        "accrued_expenses":          19,
        "short_term_debt":           20,
        "current_portion_leases":    21,
        "income_taxes_payable":      22,
        "current_unearned_revenue":  23,
        "other_current_liabilities": 24,
        "total_current_liabilities": 25,
        "long_term_debt":            26,
        "long_term_leases":          27,
        "lt_unearned_revenue":       28,
        "lt_deferred_tax_liabilities":29,
        "other_lt_liabilities":      30,
        "total_liabilities":         31,
        "common_stock":              32,
        "apic":                      33,
        "retained_earnings":         34,
        "treasury_stock":            35,
        "comprehensive_income_other":36,
        "shareholders_equity":       37,
        "total_liabilities_equity":  38,
    }

    bs_data = {}
    for year, col in bs_years:
        row_dict = {"year": year, "ticker": TICKER, "statement": "BS"}
        for field, row in BS_ROWS.items():
            row_dict[field] = v(ws_bs, row, col)
        bs_data[year] = row_dict

    # ── CF data: row 2 = [2023,2024,2025] in cols B,C,D ──────────────────
    cf_years = []
    for c in range(2, 6):
        val = ws_cf.cell(row=2, column=c).value
        if isinstance(val, int):
            cf_years.append((val, c))

    CF_ROWS = {
        "net_income":                4,
        "stock_comp":                6,
        "depreciation_amortization": 7,
        "deferred_taxes":            8,
        "gains_equity_sec":          9,
        "acq_termination":           10,
        "other_operating":           11,
        "chg_accounts_receivable":   13,
        "chg_inventory":             14,
        "chg_prepaid_other":         15,
        "chg_accounts_payable":      16,
        "chg_accrued_liabilities":   17,
        "chg_other_lt_liabilities":  18,
        "cfo":                       19,
        "proceeds_maturities_sec":   21,
        "proceeds_sales_sec":        22,
        "proceeds_sales_equity_sec": 23,
        "purchases_securities":      24,
        "capex":                     25,
        "purchases_equity_sec":      26,
        "acquisitions_net":          27,
        "other_investing":           28,
        "cfi":                       29,
        "proceeds_stock_plans":      31,
        "share_repurchases":         32,
        "tax_on_rsu":                33,
        "debt_repayment":            34,
        "dividends_paid":            35,
        "principal_lease_payments":  36,
        "other_financing":           37,
        "cff":                       38,
        "net_change_cash":           39,
        "cash_beginning":            40,
        "cash_ending":               41,
        "cash_taxes_paid":           43,
        "cash_interest_paid":        44,
    }

    cf_data = {}
    for year, col in cf_years:
        row_dict = {"year": year, "ticker": TICKER, "statement": "CF"}
        for field, row in CF_ROWS.items():
            row_dict[field] = v(ws_cf, row, col)
        cf_data[year] = row_dict

    # ── Write CSV ──────────────────────────────────────────────────────────
    all_fields_is = ["year","ticker","statement"] + list(IS_ROWS.keys())
    all_fields_bs = ["year","ticker","statement"] + list(BS_ROWS.keys())
    all_fields_cf = ["year","ticker","statement"] + list(CF_ROWS.keys())

    # Union of all fields for single file
    all_fields = sorted(set(
        list(IS_ROWS.keys()) + list(BS_ROWS.keys()) + list(CF_ROWS.keys())
    ))
    header = ["year", "ticker", "statement"] + all_fields

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        writer.writeheader()

        for year in sorted(is_data.keys()):
            writer.writerow(is_data[year])
        for year in sorted(bs_data.keys()):
            writer.writerow(bs_data[year])
        for year in sorted(cf_data.keys()):
            writer.writerow(cf_data[year])

    print(f"  ✓  {out_path}")
    print(f"       IS: {len(is_data)} years ({min(is_data)}-{max(is_data)})")
    print(f"       BS: {len(bs_data)} years ({min(bs_data)}-{max(bs_data)})")
    print(f"       CF: {len(cf_data)} years ({min(cf_data)}-{max(cf_data)})")

    return is_data, bs_data, cf_data


# ── 2. Market Data (raw) ──────────────────────────────────────────────────────

def export_market_data(wb, out_path):
    """
    Exports peer company data from 17_ComparableAnalysis sheet.
    """
    ws = wb["17_ComparableAnalysis"]

    peers = []
    peer_rows = range(6, 12)  # rows 6–11: NVIDIA + 5 peers
    col_map = {
        "company":     1,
        "price":       2,
        "eps":         3,
        "shares_b":    4,
        "mkt_cap_b":   5,
        "net_debt_b":  6,
        "ev_b":        7,
        "revenue_b":   9,
        "ebitda_b":    10,
        "net_income_b":11,
    }

    for row in peer_rows:
        rec = {"ticker": TICKER, "as_of_date": "2025-10-17"}
        for field, col in col_map.items():
            val = ws.cell(row=row, column=col).value
            rec[field] = val
        peers.append(rec)

    header = ["ticker","as_of_date","company","price","eps","shares_b",
              "mkt_cap_b","net_debt_b","ev_b","revenue_b","ebitda_b","net_income_b"]

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(peers)

    print(f"  ✓  {out_path}")
    print(f"       {len(peers)} companies (NVIDIA + 5 peers)")

    return peers


# ── 3. Cleaned Financials (processed) ─────────────────────────────────────────

def export_cleaned_financials(is_data, bs_data, cf_data, out_path):
    """
    Produces the normalized cleaned_financials.csv matching the Phase A spec:

    year, ticker, revenue, ebit, ebitda, net_income, gross_profit,
    capex, depreciation, da, change_in_nwc,
    cfo, cfi, cff,
    short_term_debt, long_term_debt, total_debt, cash,
    total_assets, total_liabilities, shareholders_equity,
    shares_outstanding,
    gross_margin, ebit_margin, net_margin, rd_pct_revenue, capex_pct_revenue
    """

    # All years that appear in at least IS data
    all_years = sorted(set(list(is_data.keys())))

    header = [
        "year", "ticker",
        # Income statement
        "revenue", "gross_profit", "ebit", "ebitda",
        "net_income", "ebt", "income_tax", "interest_income",
        "interest_expense", "rd_expense", "sga_expense", "da",
        # Cash flow
        "cfo", "cfi", "cff", "capex", "change_in_nwc",
        # Balance sheet
        "cash", "short_term_debt", "long_term_debt", "total_debt",
        "total_assets", "total_liabilities", "shareholders_equity",
        "receivables", "inventory", "accounts_payable",
        # Derived ratios
        "gross_margin", "ebit_margin", "net_margin",
        "rd_pct_revenue", "capex_pct_revenue", "fcf",
    ]

    rows = []
    for year in all_years:
        isd = is_data.get(year, {})
        bsd = bs_data.get(year, {})
        cfd = cf_data.get(year, {})

        rev   = safe(isd.get("revenue"))
        ebit  = safe(isd.get("ebit"))
        da    = safe(isd.get("da"))
        ebitda = (ebit + da) if (ebit is not None and da is not None) else None
        capex  = cfd.get("capex")   # negative in CF statement
        cfo    = cfd.get("cfo")
        fcf    = (cfo + capex) if (cfo is not None and capex is not None) else None

        # NWC change: chg_AR + chg_inventory + chg_prepaid - chg_AP - chg_accrued
        chg_ar  = cfd.get("chg_accounts_receivable")
        chg_inv = cfd.get("chg_inventory")
        chg_pre = cfd.get("chg_prepaid_other")
        chg_ap  = cfd.get("chg_accounts_payable")
        chg_acc = cfd.get("chg_accrued_liabilities")
        if all(x is not None for x in [chg_ar, chg_inv, chg_pre, chg_ap, chg_acc]):
            change_in_nwc = chg_ar + chg_inv + chg_pre + chg_ap + chg_acc
        else:
            change_in_nwc = None

        st_debt = safe(bsd.get("short_term_debt"), 0)
        lt_debt = safe(bsd.get("long_term_debt"), 0)
        lt_lease= safe(bsd.get("long_term_leases"), 0)
        cur_lease = safe(bsd.get("current_portion_leases"), 0)
        total_debt = st_debt + lt_debt + lt_lease + cur_lease if bsd else None

        row = {
            "year":       year,
            "ticker":     TICKER,
            "revenue":    rev,
            "gross_profit": safe(isd.get("gross_profit")),
            "ebit":       ebit,
            "ebitda":     round(ebitda, 2) if ebitda else None,
            "net_income": safe(isd.get("net_income")),
            "ebt":        safe(isd.get("ebt")),
            "income_tax": safe(isd.get("income_tax")),
            "interest_income":  safe(isd.get("interest_income")),
            "interest_expense": safe(isd.get("interest_expense")),
            "rd_expense":       safe(isd.get("rd_expense")),
            "sga_expense":      safe(isd.get("sga_expense")),
            "da":               safe(isd.get("da")),
            "cfo":     safe(cfd.get("cfo")),
            "cfi":     safe(cfd.get("cfi")),
            "cff":     safe(cfd.get("cff")),
            "capex":   capex,
            "change_in_nwc": round(change_in_nwc, 2) if change_in_nwc else None,
            "cash":          safe(bsd.get("cash_and_st_investments")),
            "short_term_debt": safe(bsd.get("short_term_debt")),
            "long_term_debt":  safe(bsd.get("long_term_debt")),
            "total_debt":      round(total_debt, 2) if total_debt else None,
            "total_assets":       safe(bsd.get("total_assets")),
            "total_liabilities":  safe(bsd.get("total_liabilities")),
            "shareholders_equity":safe(bsd.get("shareholders_equity")),
            "receivables":    safe(bsd.get("receivables")),
            "inventory":      safe(bsd.get("inventory")),
            "accounts_payable":safe(bsd.get("accounts_payable")),
            "gross_margin":     round(pct(safe(isd.get("gross_profit"),0), rev), 6) if rev else None,
            "ebit_margin":      round(pct(ebit, rev), 6) if (ebit and rev) else None,
            "net_margin":       round(pct(safe(isd.get("net_income"),0), rev), 6) if rev else None,
            "rd_pct_revenue":   round(pct(safe(isd.get("rd_expense"),0), rev), 6) if rev else None,
            "capex_pct_revenue":round(pct(abs(capex) if capex else 0, rev), 6) if rev else None,
            "fcf":              round(fcf, 2) if fcf else None,
        }
        rows.append(row)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    print(f"  ✓  {out_path}")
    print(f"       {len(rows)} rows | {len(header)} columns | FY{min(all_years)}–FY{max(all_years)}")

    return rows


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export datasets from Excel model")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    args = parser.parse_args()

    model_path = os.path.abspath(args.model)

    print(f"\n{'─'*58}")
    print(f"  export_datasets.py — CSV Export Pipeline")
    print(f"{'─'*58}")
    print(f"\n  Model : {model_path}\n")

    if not os.path.exists(model_path):
        print(f"  ERROR: Model not found: {model_path}\n")
        return

    print(f"  Loading workbook...")
    wb = openpyxl.load_workbook(model_path, data_only=True)

    print(f"\n  Exporting raw datasets...")
    raw_hist   = os.path.join(RAW_DIR, "nvidia_historical_financials.csv")
    raw_market = os.path.join(RAW_DIR, "nvidia_market_data.csv")
    proc_clean = os.path.join(PROCESSED_DIR, "cleaned_financials.csv")

    is_data, bs_data, cf_data = export_historical_financials(wb, raw_hist)
    peers = export_market_data(wb, raw_market)

    print(f"\n  Generating cleaned_financials.csv...")
    cleaned = export_cleaned_financials(is_data, bs_data, cf_data, proc_clean)

    # ── Print sample of cleaned output ────────────────────────────────────
    print(f"\n{'─'*58}")
    print(f"  CLEANED FINANCIALS SAMPLE  (IS rows, key metrics)")
    print(f"{'─'*58}")
    print(f"\n  {'Year':<6} {'Revenue':>10} {'EBIT':>10} "
          f"{'Net Inc':>10} {'EBIT Mgn':>10} {'Net Mgn':>10} {'FCF':>10}")
    print(f"  {'─'*4} {'─'*10} {'─'*10} {'─'*10} {'─'*10} {'─'*10} {'─'*10}")
    for row in cleaned:
        if row["statement_type"] if "statement_type" in row else True:
            rev  = row.get("revenue")
            ebit = row.get("ebit")
            ni   = row.get("net_income")
            em   = row.get("ebit_margin")
            nm   = row.get("net_margin")
            fcf  = row.get("fcf")
            if rev:
                print(
                    f"  {row['year']:<6} "
                    f"{str(rev):>10} "
                    f"{str(ebit) if ebit else 'n/a':>10} "
                    f"{str(ni) if ni else 'n/a':>10} "
                    f"{f'{em*100:.1f}%' if em else 'n/a':>10} "
                    f"{f'{nm*100:.1f}%' if nm else 'n/a':>10} "
                    f"{str(fcf) if fcf else 'n/a':>10}"
                )

    print(f"\n{'─'*58}")
    print(f"  ALL FILES WRITTEN")
    print(f"{'─'*58}")
    print(f"  datasets/raw/nvidia_historical_financials.csv")
    print(f"  datasets/raw/nvidia_market_data.csv")
    print(f"  datasets/processed/cleaned_financials.csv")
    print(f"\n  Run git add datasets/ to stage for commit.")
    print(f"{'─'*58}\n")


if __name__ == "__main__":
    main()
