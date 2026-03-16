"""
run_all.py — NVIDIA Valuation Pipeline
========================================
Master orchestrator. Runs the complete pipeline in sequence:

  Step 1  Load & validate model
  Step 2  WACC extraction + independent recalculation
  Step 3  Comparable company analysis
  Step 4  CSV dataset export
  Step 5  Write combined summary output

Usage:
    python run_all.py
    python run_all.py --model path/to/model.xlsx
    python run_all.py --skip-export     (skip CSV re-export, use existing)
    python run_all.py --verbose

Outputs:
    datasets/processed/wacc_results.json
    datasets/processed/comps_results.json
    datasets/processed/pipeline_summary.json
    datasets/raw/nvidia_historical_IS.csv
    datasets/raw/nvidia_historical_BS.csv
    datasets/raw/nvidia_historical_CF.csv
    datasets/raw/nvidia_market_data.csv
    datasets/processed/cleaned_financials.csv
"""

import os
import sys
import json
import argparse
import traceback
from datetime import datetime

import openpyxl

# ── Path setup ────────────────────────────────────────────────────────────────
ROOT          = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH    = os.path.join(ROOT, "models", "Company_Valuation_Model.xlsx")
PROCESSED_DIR = os.path.join(ROOT, "datasets", "processed")
RAW_DIR       = os.path.join(ROOT, "datasets", "raw")

WACC_OUT      = os.path.join(PROCESSED_DIR, "wacc_results.json")
COMPS_OUT     = os.path.join(PROCESSED_DIR, "comps_results.json")
SUMMARY_OUT   = os.path.join(PROCESSED_DIR, "pipeline_summary.json")

TICKER = "NVDA"


# ── Logging helpers ───────────────────────────────────────────────────────────

class Logger:
    def __init__(self, verbose=False):
        self.verbose  = verbose
        self.errors   = []
        self.warnings = []

    def info(self, msg):
        print(f"  {msg}")

    def ok(self, msg):
        print(f"  ✓  {msg}")

    def warn(self, msg):
        print(f"  ⚠  {msg}")
        self.warnings.append(msg)

    def err(self, msg):
        print(f"  ✗  {msg}")
        self.errors.append(msg)

    def detail(self, msg):
        if self.verbose:
            print(f"     {msg}")

    def section(self, title):
        print(f"\n{'─'*60}")
        print(f"  {title}")
        print(f"{'─'*60}")


log = Logger()


# ── Step 1: Model validation ──────────────────────────────────────────────────

def validate_model(model_path):
    log.section("Step 1 / 5  —  Model Validation")

    if not os.path.exists(model_path):
        log.err(f"Model not found: {model_path}")
        log.err("Place Company_Valuation_Model.xlsx in the models/ folder.")
        return None

    log.info(f"Loading: {os.path.basename(model_path)}")
    wb = openpyxl.load_workbook(model_path, data_only=True)

    # Tab check
    required_tabs = [
        "00_Assumptions", "01_Historical_IS", "02_Historical_BS",
        "03_Historical_CF", "05b_DCF", "15_WACC", "05a_FCFF",
        "17_ComparableAnalysis"
    ]
    missing = [t for t in required_tabs if t not in wb.sheetnames]
    if missing:
        log.warn(f"Missing expected tabs: {missing}")
    else:
        log.ok(f"All required tabs present ({len(wb.sheetnames)} total)")

    # Named ranges
    named = list(wb.defined_names)
    expected_ranges = ["risk_free_rate", "adjusted_beta",
                       "equity_risk_premium", "effective_tax_rate", "wacc_output"]
    missing_ranges = [r for r in expected_ranges if r not in named]
    if missing_ranges:
        log.warn(f"Missing named ranges: {missing_ranges}")
        log.warn("Run u10_fix.py to add named ranges.")
    else:
        log.ok(f"All 5 named ranges present")

    # Critical output cell
    dcf_price = wb["05b_DCF"]["C21"].value
    if dcf_price and isinstance(dcf_price, float):
        log.ok(f"DCF implied price: ${dcf_price:.2f}")
    else:
        log.warn("DCF!C21 is None — open model in Excel, Ctrl+Alt+F9, Save")

    return wb


# ── Step 2: WACC ──────────────────────────────────────────────────────────────

def run_wacc(wb):
    log.section("Step 2 / 5  —  WACC Extraction & Recalculation")

    import statistics as _stats

    NAMED = {
        "risk_free_rate":      ("15_WACC", "F10", 0.0407),
        "adjusted_beta":       ("15_WACC", "F11", 1.7728),
        "equity_risk_premium": ("15_WACC", "F12", 0.05),
        "effective_tax_rate":  ("15_WACC", "F17", 0.1326),
        "wacc_output":         ("15_WACC", "F19", 0.1291),
    }
    ADDR = {
        "market_cap_equity":   ("15_WACC", "F7"),
        "weight_equity":       ("15_WACC", "F8"),
        "total_debt":          ("15_WACC", "F14"),
        "weight_debt":         ("15_WACC", "F15"),
        "pretax_cost_of_debt": ("15_WACC", "F16"),
    }

    inputs = {}

    # Read named ranges
    for name, (sheet, addr, expected) in NAMED.items():
        val = wb[sheet][addr].value
        if val is None:
            log.warn(f"  {name} = None (formula cell needs Excel recalculation)")
            inputs[name] = expected   # fall back to known value
        else:
            inputs[name] = float(val)
            log.detail(f"{name} = {val}")

    # Read address inputs
    for name, (sheet, addr) in ADDR.items():
        val = wb[sheet][addr].value
        if isinstance(val, (int, float)):
            inputs[name] = float(val)
            log.detail(f"{name} = {val}")
        else:
            log.warn(f"{name} at {sheet}!{addr} not numeric: {repr(val)}")
            inputs[name] = None

    # Recalculate
    rf   = inputs["risk_free_rate"]
    beta = inputs["adjusted_beta"]
    erp  = inputs["equity_risk_premium"]
    tax  = inputs["effective_tax_rate"]
    w_eq = inputs.get("weight_equity") or 0.9981
    kd   = inputs.get("pretax_cost_of_debt") or 0.0283
    w_d  = inputs.get("weight_debt") or 0.0019

    ke        = rf + beta * erp
    kd_post   = kd * (1 - tax)
    wacc_calc = w_eq * ke + w_d * kd_post
    wacc_mdl  = inputs["wacc_output"]
    diff_bp   = abs(wacc_calc - wacc_mdl) * 10000

    log.ok(f"Cost of equity (CAPM):  {ke*100:.4f}%")
    log.ok(f"After-tax cost of debt: {kd_post*100:.4f}%")
    log.ok(f"WACC recalculated:      {wacc_calc*100:.4f}%")
    log.ok(f"WACC from model:        {wacc_mdl*100:.4f}%")

    if diff_bp < 1:
        log.ok(f"Cross-check: PASS ({diff_bp:.2f} bps difference)")
    else:
        log.warn(f"Cross-check: {diff_bp:.2f} bps difference — investigate")

    result = {
        "meta": {
            "step": "wacc",
            "run_timestamp": datetime.now().isoformat(),
            "cross_check": "PASS" if diff_bp < 1 else "FAIL",
            "diff_basis_pts": round(diff_bp, 4),
        },
        "inputs": {
            "risk_free_rate": rf, "adjusted_beta": beta,
            "equity_risk_premium": erp, "effective_tax_rate": tax,
            "weight_equity": w_eq, "weight_debt": w_d,
            "pretax_cost_of_debt": kd,
            "market_cap_equity_m": inputs.get("market_cap_equity"),
            "total_debt_m": inputs.get("total_debt"),
        },
        "outputs": {
            "cost_of_equity":          round(ke, 8),
            "after_tax_cost_of_debt":  round(kd_post, 8),
            "wacc_recalculated":       round(wacc_calc, 8),
            "wacc_from_model":         round(wacc_mdl, 8),
        },
        "display": {
            "cost_of_equity_pct":          f"{ke*100:.4f}%",
            "after_tax_cost_of_debt_pct":  f"{kd_post*100:.4f}%",
            "wacc_recalculated_pct":       f"{wacc_calc*100:.4f}%",
            "wacc_from_model_pct":         f"{wacc_mdl*100:.4f}%",
        }
    }

    os.makedirs(PROCESSED_DIR, exist_ok=True)
    with open(WACC_OUT, "w") as f:
        json.dump(result, f, indent=2)
    log.ok(f"Saved: datasets/processed/wacc_results.json")

    return result


# ── Step 3: Comps ─────────────────────────────────────────────────────────────

def run_comps(wb):
    log.section("Step 3 / 5  —  Comparable Company Analysis")

    import statistics as _stats

    ws = wb["17_ComparableAnalysis"]

    COL = {"company":1,"price":2,"eps":3,"shares_b":4,"mkt_cap_b":5,
           "net_debt_b":6,"ev_b":7,"revenue_b":9,"ebitda_b":10,"net_income_b":11}

    def read_row(row):
        rec = {}
        for field, col in COL.items():
            v = ws.cell(row=row, column=col).value
            rec[field] = float(v) if isinstance(v, (int, float)) else None
        rec["company"] = ws.cell(row=row, column=1).value
        return rec

    subject = read_row(6)
    peers   = [read_row(r) for r in range(7, 12)]

    def multiples(p):
        ev_rev  = (p["ev_b"] / p["revenue_b"])  if p["ev_b"] and p["revenue_b"] and p["revenue_b"] > 0 else None
        ev_ebit = (p["ev_b"] / p["ebitda_b"])   if p["ev_b"] and p["ebitda_b"] and p["ebitda_b"]  > 0 else None
        pe      = (p["price"] / p["eps"])        if p["price"] and p["eps"] and p["eps"] > 0 else None
        return {**p, "ev_revenue": ev_rev, "ev_ebitda": ev_ebit, "pe": pe}

    peer_mults = [multiples(p) for p in peers]

    def stats(vals):
        clean = sorted([v for v in vals if v is not None])
        if not clean: return {}
        n = len(clean)
        def pct(data, p):
            idx = (p/100)*(len(data)-1)
            lo, hi = int(idx), min(int(idx)+1, len(data)-1)
            return data[lo] + (idx-lo)*(data[hi]-data[lo])
        return {"high": max(clean), "p75": pct(clean,75), "mean": sum(clean)/n,
                "median": _stats.median(clean), "p25": pct(clean,25), "low": min(clean), "n": n}

    stat_evr  = stats([p["ev_revenue"] for p in peer_mults])
    stat_eve  = stats([p["ev_ebitda"]  for p in peer_mults])
    stat_pe   = stats([p["pe"]         for p in peer_mults])

    price  = subject["price"]
    shares = subject["shares_b"]
    nd     = subject["net_debt_b"]
    rev    = subject["revenue_b"]
    ebitda = subject["ebitda_b"]
    eps    = subject["eps"]

    def imp_ev(mult, metric):
        if mult is None or metric is None or shares is None: return None
        return (mult * metric - nd) / shares

    def imp_pe(mult):
        return (mult * eps) if mult and eps else None

    implied = {}
    for stat in ["high","p75","mean","median","p25","low"]:
        implied[stat] = {
            "ev_revenue": round(imp_ev(stat_evr.get(stat), rev), 2)  if stat_evr.get(stat) else None,
            "ev_ebitda":  round(imp_ev(stat_eve.get(stat), ebitda), 2) if stat_eve.get(stat) else None,
            "pe":         round(imp_pe(stat_pe.get(stat)), 2)          if stat_pe.get(stat)  else None,
        }

    med = implied["median"]
    log.ok(f"Peers: {[p['company'] for p in peer_mults]}")
    log.ok(f"Median EV/Revenue:  {stat_evr.get('median',0):.1f}x  → implied ${med['ev_revenue']}")
    log.ok(f"Median EV/EBITDA:   {stat_eve.get('median',0):.1f}x  → implied ${med['ev_ebitda']}")
    log.ok(f"Median P/E:         {stat_pe.get('median',0):.1f}x   → implied ${med['pe']}")
    log.ok(f"Current price: ${price:.2f}")

    def ud(implied_p):
        if implied_p and price:
            chg = (implied_p/price - 1)*100
            return f"{chg:+.1f}%"
        return "n/a"
    log.ok(f"EV/EBITDA median vs market: {ud(med['ev_ebitda'])}")

    result = {
        "meta": {"step": "comps", "run_timestamp": datetime.now().isoformat(),
                 "subject": subject["company"], "n_peers": len(peer_mults),
                 "peer_universe": [p["company"] for p in peer_mults]},
        "subject_inputs": {"current_price": price, "eps": eps, "shares_b": shares,
                           "net_debt_b": nd, "revenue_b": rev, "ebitda_b": ebitda},
        "peer_multiples": [{"company": p["company"], "ev_revenue": round(p["ev_revenue"],3) if p["ev_revenue"] else None,
                             "ev_ebitda": round(p["ev_ebitda"],3) if p["ev_ebitda"] else None,
                             "pe": round(p["pe"],2) if p["pe"] else None} for p in peer_mults],
        "multiple_statistics": {"ev_revenue": stat_evr, "ev_ebitda": stat_eve, "pe": stat_pe},
        "implied_prices": implied,
        "summary": {"current_price": price,
                    "median_implied_ev_revenue": med["ev_revenue"],
                    "median_implied_ev_ebitda":  med["ev_ebitda"],
                    "median_implied_pe":         med["pe"]},
    }

    with open(COMPS_OUT, "w") as f:
        json.dump(result, f, indent=2)
    log.ok(f"Saved: datasets/processed/comps_results.json")

    return result


# ── Step 4: CSV export ────────────────────────────────────────────────────────

def run_export(wb):
    log.section("Step 4 / 5  —  Dataset Export")

    import csv

    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    def v(ws, row, col):
        val = ws.cell(row=row, column=col).value
        return val if isinstance(val, (int, float)) else None

    # ── IS export ─────────────────────────────────────────────────────────
    ws_is = wb["01_Historical_IS"]
    ws_p  = wb["04a_Projection_IS"]
    ws_gr = wb["10_GrowthRates"]
    ws_cf = wb["03_Historical_CF"]

    is_year_col = {}
    for c in range(2, 9):
        val = ws_is.cell(row=2, column=c).value
        if isinstance(val, int) and 2015 < val < 2026:
            is_year_col[val] = c

    da_is = {2020: v(ws_is,13,2), 2021: v(ws_is,13,3), 2022: v(ws_is,13,4)}
    cf_yc = {}
    for c in range(2, 6):
        yr = ws_cf.cell(row=2, column=c).value
        if isinstance(yr, int): cf_yc[yr] = c
    da_cf = {yr: v(ws_cf,7,col) for yr,col in cf_yc.items()}
    da_all = {**da_is, **da_cf}

    gr_yc = {}
    for c in range(3, 15):
        yr = v(ws_gr, 1, c)
        if yr: gr_yc[int(yr)] = c
    rd_gr = {yr: v(ws_gr,6,col) for yr,col in gr_yc.items() if 2019 < yr < 2023}

    IS_ROWS = {'revenue':3,'cogs':4,'gross_profit':5,'rd_expense':8,'sga_expense':9,
               'acq_termination':10,'total_opex':11,'ebit':12,'interest_income':15,
               'interest_expense':16,'other_net':17,'ebt':19,'income_tax':20,'net_income':21}
    IS_FIELDS = ['year','ticker','revenue','cogs','gross_profit','rd_expense','sga_expense',
                 'acq_termination','total_opex','ebit','da','ebitda','interest_income',
                 'interest_expense','other_net','ebt','income_tax','net_income']

    is_rows = []
    for year, col in is_year_col.items():
        d = {'year': year, 'ticker': TICKER}
        for field, row in IS_ROWS.items():
            d[field] = v(ws_is, row, col)
        if d['gross_profit'] is None and d['revenue'] and d['cogs']:
            d['gross_profit'] = d['revenue'] - d['cogs']
        if d['rd_expense'] is None and year in rd_gr:
            d['rd_expense'] = rd_gr[year]
        if d['ebit'] is None and d.get('gross_profit') is not None:
            d['ebit'] = d['gross_profit'] - (d['rd_expense'] or 0) - (d['sga_expense'] or 0) - (d['acq_termination'] or 0)
        d['da'] = da_all.get(year)
        d['ebitda'] = (d['ebit'] + d['da']) if (d['ebit'] and d['da']) else None
        is_rows.append(d)

    path_is = os.path.join(RAW_DIR, "nvidia_historical_IS.csv")
    with open(path_is, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=IS_FIELDS, extrasaction='ignore')
        w.writeheader(); w.writerows(is_rows)
    log.ok(f"IS: {path_is.split('datasets/')[-1]}  ({len(is_rows)} years)")

    # ── BS export ─────────────────────────────────────────────────────────
    ws_bs = wb["02_Historical_BS"]
    bs_yc = {}
    for c in range(2, 9):
        val = ws_bs.cell(row=2, column=c).value
        if isinstance(val, str) and val.strip().isdigit():
            bs_yc[int(val.strip())] = c
        elif isinstance(val, int) and 2018 < val < 2026:
            bs_yc[val] = c

    BS_ROWS = {'cash_equivalents':4,'short_term_investments':5,'cash_and_st_investments':6,
               'receivables':7,'inventory':8,'prepaid_expenses':9,'total_current_assets':10,
               'ppe_net':11,'operating_lease_assets':12,'goodwill':13,'intangible_assets':14,
               'lt_deferred_tax_assets':15,'other_assets':16,'total_assets':17,
               'accounts_payable':18,'accrued_expenses':19,'short_term_debt':20,
               'current_portion_leases':21,'income_taxes_payable':22,'current_unearned_revenue':23,
               'other_current_liabilities':24,'total_current_liabilities':25,'long_term_debt':26,
               'long_term_leases':27,'lt_unearned_revenue':28,'lt_deferred_tax_liabilities':29,
               'other_lt_liabilities':30,'total_liabilities':31,'common_stock':32,'apic':33,
               'retained_earnings':34,'treasury_stock':35,'comprehensive_income_other':36,
               'shareholders_equity':37,'total_liabilities_equity':38}
    BS_FIELDS = ['year','ticker'] + list(BS_ROWS.keys())

    bs_rows = []
    for year, col in bs_yc.items():
        d = {'year': year, 'ticker': TICKER}
        for field, row in BS_ROWS.items():
            d[field] = v(ws_bs, row, col)
        bs_rows.append(d)

    path_bs = os.path.join(RAW_DIR, "nvidia_historical_BS.csv")
    with open(path_bs, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=BS_FIELDS, extrasaction='ignore')
        w.writeheader(); w.writerows(bs_rows)
    log.ok(f"BS: {path_bs.split('datasets/')[-1]}  ({len(bs_rows)} years)")

    # ── CF export ─────────────────────────────────────────────────────────
    CF_ROWS = {'net_income':4,'stock_comp':6,'depreciation_amortization':7,'deferred_taxes':8,
               'gains_equity_sec':9,'acq_termination':10,'other_operating':11,
               'chg_accounts_receivable':13,'chg_inventory':14,'chg_prepaid_other':15,
               'chg_accounts_payable':16,'chg_accrued_liabilities':17,'chg_other_lt_liabilities':18,
               'cfo':19,'proceeds_maturities_sec':21,'proceeds_sales_sec':22,
               'proceeds_sales_equity_sec':23,'purchases_securities':24,'capex':25,
               'purchases_equity_sec':26,'acquisitions_net':27,'other_investing':28,'cfi':29,
               'proceeds_stock_plans':31,'share_repurchases':32,'tax_on_rsu':33,
               'debt_repayment':34,'dividends_paid':35,'principal_lease_payments':36,
               'other_financing':37,'cff':38,'net_change_cash':39,'cash_beginning':40,
               'cash_ending':41,'cash_taxes_paid':43,'cash_interest_paid':44}
    CF_FIELDS = ['year','ticker'] + list(CF_ROWS.keys()) + ['fcf']

    cf_rows = []
    # FY2022 from secondary section (col C = FY2022)
    fy22_map = {'net_income':205,'depreciation_amortization':206,'gains_equity_sec':207,
                'stock_comp':208,'other_operating':209,'chg_accounts_receivable':210,
                'chg_inventory':211,'chg_accounts_payable':212,'chg_other_lt_liabilities':213,
                'cfo':214,'capex':216,'acquisitions_net':217,'purchases_securities':218,
                'other_investing':219,'cfi':220,'debt_repayment':222,'proceeds_stock_plans':224,
                'share_repurchases':225,'dividends_paid':226,'other_financing':227,'cff':228,
                'net_change_cash':229,'cash_interest_paid':234,'cash_taxes_paid':235}
    d22 = {'year': 2022, 'ticker': TICKER}
    for field in CF_ROWS: d22[field] = None
    for field, row in fy22_map.items():
        d22[field] = v(ws_cf, row, 3)
    d22['fcf'] = (d22['cfo'] + d22['capex']) if (d22['cfo'] and d22['capex']) else None
    cf_rows.append(d22)

    for year, col in cf_yc.items():
        d = {'year': year, 'ticker': TICKER}
        for field, row in CF_ROWS.items():
            d[field] = v(ws_cf, row, col)
        cfo = d.get('cfo'); cap = d.get('capex')
        d['fcf'] = round(cfo + cap, 2) if (cfo and cap) else None
        cf_rows.append(d)

    path_cf = os.path.join(RAW_DIR, "nvidia_historical_CF.csv")
    with open(path_cf, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=CF_FIELDS, extrasaction='ignore')
        w.writeheader(); w.writerows(sorted(cf_rows, key=lambda x: x['year']))
    log.ok(f"CF: {path_cf.split('datasets/')[-1]}  ({len(cf_rows)} years)")

    # ── Market data ───────────────────────────────────────────────────────
    ws_c = wb["17_ComparableAnalysis"]
    mkt = []
    for row in range(6, 12):
        rec = {'ticker': TICKER, 'as_of_date': '2025-10-17'}
        for field, col in {'company':1,'price':2,'eps':3,'shares_b':4,'mkt_cap_b':5,
                           'net_debt_b':6,'ev_b':7,'revenue_b':9,'ebitda_b':10,'net_income_b':11}.items():
            rec[field] = ws_c.cell(row=row, column=col).value
        mkt.append(rec)
    path_mkt = os.path.join(RAW_DIR, "nvidia_market_data.csv")
    mkt_fields = ['ticker','as_of_date','company','price','eps','shares_b','mkt_cap_b',
                  'net_debt_b','ev_b','revenue_b','ebitda_b','net_income_b']
    with open(path_mkt, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=mkt_fields, extrasaction='ignore')
        w.writeheader(); w.writerows(mkt)
    log.ok(f"Market data: {path_mkt.split('datasets/')[-1]}  ({len(mkt)} companies)")

    # ── cleaned_financials ────────────────────────────────────────────────
    is_d = {d['year']: d for d in is_rows}
    bs_d = {d['year']: d for d in bs_rows}
    cf_d = {d['year']: d for d in cf_rows}

    def pct(a, b):
        return round(a/b, 6) if (a is not None and b and b != 0) else None

    cleaned = []
    for year in sorted(set(list(is_d)+list(bs_d)+list(cf_d))):
        isd = is_d.get(year, {})
        bsd = bs_d.get(year, {})
        cfd = cf_d.get(year, {})
        rev = isd.get('revenue'); ebit = isd.get('ebit'); ni = isd.get('net_income')
        cfo = cfd.get('cfo');     cap  = cfd.get('capex')
        fcf = (cfo + cap) if (cfo and cap) else None
        chg_vals = [cfd.get(f) for f in ['chg_accounts_receivable','chg_inventory',
                    'chg_prepaid_other','chg_accounts_payable','chg_accrued_liabilities']]
        nwc = sum(chg_vals) if all(x is not None for x in chg_vals) else None
        td  = sum(filter(None,[bsd.get('short_term_debt',0),bsd.get('long_term_debt',0),
                               bsd.get('long_term_leases',0),bsd.get('current_portion_leases',0)]))
        cleaned.append({
            'year':year,'ticker':TICKER,
            'revenue':rev,'gross_profit':isd.get('gross_profit'),'ebit':ebit,
            'ebitda':isd.get('ebitda'),'net_income':ni,'ebt':isd.get('ebt'),
            'income_tax':isd.get('income_tax'),'interest_income':isd.get('interest_income'),
            'interest_expense':isd.get('interest_expense'),'rd_expense':isd.get('rd_expense'),
            'sga_expense':isd.get('sga_expense'),'da':isd.get('da'),
            'cfo':cfo,'cfi':cfd.get('cfi'),'cff':cfd.get('cff'),'capex':cap,
            'change_in_nwc': round(nwc,2) if nwc else None,
            'cash':bsd.get('cash_and_st_investments'),
            'short_term_debt':bsd.get('short_term_debt'),'long_term_debt':bsd.get('long_term_debt'),
            'total_debt': round(td,2) if td else None,
            'total_assets':bsd.get('total_assets'),'total_liabilities':bsd.get('total_liabilities'),
            'shareholders_equity':bsd.get('shareholders_equity'),
            'receivables':bsd.get('receivables'),'inventory':bsd.get('inventory'),
            'accounts_payable':bsd.get('accounts_payable'),
            'gross_margin':pct(isd.get('gross_profit'),rev),'ebit_margin':pct(ebit,rev),
            'net_margin':pct(ni,rev),'rd_pct_revenue':pct(isd.get('rd_expense'),rev),
            'capex_pct_revenue':pct(abs(cap) if cap else 0, rev),
            'fcf': round(fcf,2) if fcf else None,
        })

    cln_fields = ['year','ticker','revenue','gross_profit','ebit','ebitda','net_income',
                  'ebt','income_tax','interest_income','interest_expense','rd_expense',
                  'sga_expense','da','cfo','cfi','cff','capex','change_in_nwc','cash',
                  'short_term_debt','long_term_debt','total_debt','total_assets',
                  'total_liabilities','shareholders_equity','receivables','inventory',
                  'accounts_payable','gross_margin','ebit_margin','net_margin',
                  'rd_pct_revenue','capex_pct_revenue','fcf']
    path_cln = os.path.join(PROCESSED_DIR, "cleaned_financials.csv")
    with open(path_cln, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=cln_fields, extrasaction='ignore')
        w.writeheader(); w.writerows(cleaned)
    log.ok(f"Cleaned: {path_cln.split('datasets/')[-1]}  ({len(cleaned)} rows, {len(cln_fields)} cols)")

    return {"is_years": len(is_rows), "bs_years": len(bs_rows),
            "cf_years": len(cf_rows), "cleaned_rows": len(cleaned)}


# ── Step 5: comps_data.csv (Phase B format) ───────────────────────────────────

def run_comps_csv(wb):
    log.section("Step 5 / 5  —  comps/comps_data.csv (Phase B format)")
    import csv

    ws = wb["17_ComparableAnalysis"]

    COMPS_DIR = os.path.join(ROOT, "comps")
    os.makedirs(COMPS_DIR, exist_ok=True)

    COL = {"company":1,"price":2,"eps":3,"shares_b":4,"mkt_cap_b":5,
           "net_debt_b":6,"ev_b":7,"revenue_b":9,"ebitda_b":10,"net_income_b":11}

    # Notes per peer — adjustments and selection rationale
    peer_notes = {
        "NVIDIA":          "Subject company. Fabless GPU/AI accelerator leader.",
        "AMD":             "Peer: CPU+GPU overlap; high growth; comparable AI/data center exposure.",
        "Intel":           "Peer: CPU/semiconductor; mature; P/E excluded (net loss FY2024). Impairment charges in EBITDA.",
        "Qualcomm (QCOM)": "Peer: fabless semiconductor; mobile/automotive/IoT; lower AI exposure than NVDA.",
        "Broadcom (AVGO)": "Peer: ASIC/networking chips; data center overlap; high EBITDA margin.",
        "TSMC":            "Peer: foundry (not fabless); included for semiconductor scale comparison. Business model differs.",
    }

    rows = []
    for row in range(6, 12):
        company = ws.cell(row=row, column=1).value or ""
        price   = ws.cell(row=row, column=2).value
        eps     = ws.cell(row=row, column=3).value
        shares  = ws.cell(row=row, column=4).value
        mktcap  = ws.cell(row=row, column=5).value
        net_dbt = ws.cell(row=row, column=6).value
        ev      = ws.cell(row=row, column=7).value
        rev     = ws.cell(row=row, column=9).value
        ebitda  = ws.cell(row=row, column=10).value
        ni      = ws.cell(row=row, column=11).value

        # Compute EV from components if available
        if mktcap and net_dbt is not None:
            ev_computed = mktcap + net_dbt
        else:
            ev_computed = ev

        # Multiples
        ev_ebitda = round(ev/ebitda, 2)  if (ev and ebitda and ebitda > 0) else None
        pe        = round(price/eps, 2)  if (price and eps and eps > 0)    else None
        ev_sales  = round(ev/rev, 2)     if (ev and rev and rev > 0)       else None

        # EBITDA adj (documented adjustments)
        ebitda_adj = ebitda  # no material adjustments for NVDA/AMD/QCOM/AVGO/TSMC at this level
        # Intel: FY2024 impairment ~$15.9B noted but not added back (conservative)

        # Ticker cleanup
        ticker_map = {"NVIDIA":"NVDA","AMD":"AMD","Intel":"INTC",
                      "Qualcomm (QCOM)":"QCOM","Broadcom (AVGO)":"AVGO","TSMC":"TSM"}
        ticker = ticker_map.get(company, company[:4].upper())

        rows.append({
            "ticker":           ticker,
            "company_name":     company,
            "market_cap":       mktcap,
            "net_debt":         net_dbt,
            "ev":               ev_computed,
            "ltm_revenue":      rev,
            "ltm_ebitda":       ebitda,
            "ltm_ebitda_adj":   ebitda_adj,
            "ltm_net_income":   ni,
            "ev_ebitda":        ev_ebitda,
            "ev_ebitda_adj":    round(ev/ebitda_adj, 2) if (ev and ebitda_adj and ebitda_adj > 0) else None,
            "p_e":              pe,
            "ev_sales":         ev_sales,
            "ltm_price":        price,
            "ltm_eps":          eps,
            "data_date":        "2025-10-17",
            "notes":            peer_notes.get(company, ""),
        })

    fields = ["ticker","company_name","market_cap","net_debt","ev","ltm_revenue",
              "ltm_ebitda","ltm_ebitda_adj","ltm_net_income","ev_ebitda","ev_ebitda_adj",
              "p_e","ev_sales","ltm_price","ltm_eps","data_date","notes"]

    path = os.path.join(COMPS_DIR, "comps_data.csv")
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader(); w.writerows(rows)

    log.ok(f"Saved: comps/comps_data.csv  ({len(rows)} companies, {len(fields)} fields)")

    # Print table
    print()
    print(f"    {'Ticker':<6} {'EV/Rev':>8} {'EV/EBITDA':>10} {'P/E':>8}  Company")
    print(f"    {'─'*4} {'─'*8} {'─'*10} {'─'*8}  {'─'*25}")
    for r in rows:
        ev_r = f"{r['ev_sales']:.1f}x"    if r['ev_sales']  else "  n/m"
        ev_e = f"{r['ev_ebitda']:.1f}x"   if r['ev_ebitda'] else "  n/m"
        pe   = f"{r['p_e']:.1f}x"         if r['p_e']       else "  n/m"
        print(f"    {r['ticker']:<6} {ev_r:>8} {ev_e:>10} {pe:>8}  {r['company_name']}")

    return {"path": path, "rows": len(rows)}


# ── Summary ───────────────────────────────────────────────────────────────────

def write_summary(wacc_r, comps_r, export_r, comps_csv_r, model_path, args):
    summary = {
        "pipeline": "NVIDIA DCF Valuation — run_all.py",
        "run_timestamp": datetime.now().isoformat(),
        "model_file": os.path.basename(model_path),
        "steps_completed": ["validate", "wacc", "comps", "export", "comps_csv"],
        "warnings": log.warnings,
        "errors":   log.errors,
        "wacc_summary": {
            "wacc_pct":        wacc_r["display"]["wacc_recalculated_pct"],
            "cross_check":     wacc_r["meta"]["cross_check"],
            "diff_basis_pts":  wacc_r["meta"]["diff_basis_pts"],
        },
        "comps_summary": {
            "n_peers":                    comps_r["meta"]["n_peers"],
            "median_implied_ev_revenue":  comps_r["summary"]["median_implied_ev_revenue"],
            "median_implied_ev_ebitda":   comps_r["summary"]["median_implied_ev_ebitda"],
            "median_implied_pe":          comps_r["summary"]["median_implied_pe"],
            "current_price":              comps_r["summary"]["current_price"],
        },
        "datasets_exported": export_r,
        "comps_csv": comps_csv_r,
        "outputs": [
            "datasets/processed/wacc_results.json",
            "datasets/processed/comps_results.json",
            "datasets/processed/cleaned_financials.csv",
            "datasets/processed/pipeline_summary.json",
            "datasets/raw/nvidia_historical_IS.csv",
            "datasets/raw/nvidia_historical_BS.csv",
            "datasets/raw/nvidia_historical_CF.csv",
            "datasets/raw/nvidia_market_data.csv",
            "comps/comps_data.csv",
        ]
    }
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    with open(SUMMARY_OUT, "w") as f:
        json.dump(summary, f, indent=2)
    return summary


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="NVIDIA Valuation Pipeline")
    parser.add_argument("--model",       default=MODEL_PATH)
    parser.add_argument("--skip-export", action="store_true",
                        help="Skip CSV re-export (use existing CSVs)")
    parser.add_argument("--verbose",     action="store_true")
    args = parser.parse_args()

    log.verbose = args.verbose
    model_path  = os.path.abspath(args.model)

    print(f"\n{'═'*60}")
    print(f"  NVIDIA Valuation Pipeline  —  run_all.py")
    print(f"  {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}")
    print(f"{'═'*60}")

    # ── Run all steps ──────────────────────────────────────────────────────
    wb = validate_model(model_path)
    if wb is None:
        print(f"\n  Pipeline aborted — model not found.\n")
        sys.exit(1)

    wacc_r     = run_wacc(wb)
    comps_r    = run_comps(wb)
    export_r   = run_export(wb) if not args.skip_export else {"skipped": True}
    comps_csv_r= run_comps_csv(wb)
    summary    = write_summary(wacc_r, comps_r, export_r, comps_csv_r, model_path, args)

    # ── Final print ────────────────────────────────────────────────────────
    print(f"\n{'═'*60}")
    print(f"  PIPELINE COMPLETE")
    print(f"{'═'*60}")
    print(f"\n  WACC          {summary['wacc_summary']['wacc_pct']}  "
          f"(cross-check: {summary['wacc_summary']['cross_check']}, "
          f"{summary['wacc_summary']['diff_basis_pts']:.2f} bps)")
    cs = summary["comps_summary"]
    print(f"\n  Implied prices at peer median multiples:")
    print(f"    EV/Revenue  →  ${cs['median_implied_ev_revenue']}")
    print(f"    EV/EBITDA   →  ${cs['median_implied_ev_ebitda']}")
    print(f"    P/E         →  ${cs['median_implied_pe']}")
    print(f"    Market price   ${cs['current_price']}")

    if log.warnings:
        print(f"\n  Warnings ({len(log.warnings)}):")
        for w in log.warnings: print(f"    ⚠  {w}")
    if log.errors:
        print(f"\n  Errors ({len(log.errors)}):")
        for e in log.errors: print(f"    ✗  {e}")
        sys.exit(1)

    print(f"\n  9 output files written to datasets/ and comps/")
    print(f"  See datasets/processed/pipeline_summary.json for full log.")
    print(f"\n{'═'*60}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
